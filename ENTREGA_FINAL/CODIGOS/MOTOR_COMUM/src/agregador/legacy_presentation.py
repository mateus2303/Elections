"""Saídas essenciais do Modelo A, no formato do agregador legado.

Este módulo é intencionalmente pequeno: ele reproduz as três peças macro que
existiam no processo anterior, sem criar painel adicional, recortes
demográficos ou gráficos de primeiro turno.  Os cálculos usam os mesmos pesos
do ``aggregate_model_a``: qualidade do instituto, recência e tamanho da
amostra, todos com expoentes configuráveis.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from .configuration import Config
from .domain import ConfigurationError, is_present, normalized_text
from .standardization import StandardizedData
from .weighting import _sample_used


Z_95 = 1.959963984540054
GOVERNMENT_OUTPUT = "01_Avaliacao_do_Governo"
NET_GOVERNMENT_OUTPUT = "02_Avaliacao_Liquida_do_Governo"
SECOND_ROUND_OUTPUT = "03_Segundo_Turno_Lula_vs_Flavio"


@dataclass(frozen=True)
class SmoothedSeries:
    """Estimativa e intervalo amostral para uma série temporal."""

    values: pd.DataFrame
    raw: pd.DataFrame


def _total_brasil(frame: pd.DataFrame) -> pd.DataFrame:
    """Mantém somente o resultado nacional total publicado."""
    if frame.empty:
        return frame.copy()
    required = {"nivel_geografico", "geografia", "segmento_tipo", "segmento"}
    missing = required.difference(frame.columns)
    if missing:
        raise ConfigurationError(f"Não foi possível identificar o total Brasil: faltam {sorted(missing)}.")
    mask = (
        frame["nivel_geografico"].map(normalized_text).eq("brasil")
        & frame["geografia"].map(normalized_text).eq("brasil")
        & frame["segmento_tipo"].map(normalized_text).eq("total")
        & frame["segmento"].map(normalized_text).eq("total")
    )
    return frame.loc[mask].copy()


def _prepared(frame: pd.DataFrame, value_columns: Iterable[str], config: Config) -> pd.DataFrame:
    """Prepara componentes de peso sem alterar a base padronizada."""
    result = frame.copy()
    result["data_referencia"] = pd.to_datetime(result["data_referencia"], errors="coerce").dt.normalize()
    result = result.loc[result["data_referencia"].notna()].copy()
    for column in value_columns:
        result[column] = pd.to_numeric(result[column], errors="coerce")
    if result.empty:
        return result
    samples = result.apply(lambda row: _sample_used(row, config), axis=1)
    result["_amostra_usada"] = [item[0] for item in samples]
    result["_qualidade"] = pd.to_numeric(result["peso_legacy"], errors="coerce").fillna(
        float(config.value("unknown_institute_weight"))
    )
    result["_amostra_relativa"] = np.minimum(
        result["_amostra_usada"].astype(float), float(config.value("sample_cap"))
    ) / float(config.value("sample_cap"))
    return result


def _smooth(
    frame: pd.DataFrame,
    value_column: str,
    config: Config,
    window_days: int,
    variance_columns: tuple[str, ...] | None = None,
    limits: tuple[float, float] | None = (0.0, 100.0),
) -> SmoothedSeries:
    """Replica a suavização geométrica e o IC amostral do legado."""
    columns = [value_column, *(variance_columns or (value_column,))]
    prepared = _prepared(frame, columns, config)
    raw = prepared.loc[prepared[value_column].notna(), ["data_referencia", value_column]].copy()
    raw = raw.rename(columns={value_column: "valor_pct"}).sort_values("data_referencia")
    if raw.empty:
        return SmoothedSeries(pd.DataFrame(columns=["estimativa_pct", "ic_baixo_pct", "ic_alto_pct", "se_pp"]), raw)

    valid = prepared.loc[prepared[value_column].notna()].copy()
    dates = valid["data_referencia"].to_numpy(dtype="datetime64[D]").astype("int64")
    reference_dates = np.unique(dates)
    values = valid[value_column].to_numpy(dtype=float)
    quality = valid["_qualidade"].to_numpy(dtype=float)
    relative_sample = valid["_amostra_relativa"].to_numpy(dtype=float)
    sample_used = valid["_amostra_usada"].to_numpy(dtype=float)
    variance = np.zeros(len(valid), dtype=float)
    for column in variance_columns or (value_column,):
        proportions = valid[column].to_numpy(dtype=float) / 100
        variance += proportions * (1 - proportions)

    weight_config = config.nested("weights")
    records: list[dict[str, float | pd.Timestamp]] = []
    for reference in reference_dates:
        ages = reference - dates
        selected = np.flatnonzero((ages >= 0) & (ages <= window_days))
        if not len(selected):
            continue
        time_component = np.maximum(0.0, 1 - ages[selected] / window_days)
        raw_weight = (
            quality[selected] ** float(weight_config["quality"])
            * time_component ** float(weight_config["recency"])
            * relative_sample[selected] ** float(weight_config["sample"])
        )
        positive = raw_weight > 0
        selected, raw_weight = selected[positive], raw_weight[positive]
        if not len(selected):
            continue
        normalized_weight = raw_weight / raw_weight.sum()
        estimate = float(np.sum(normalized_weight * values[selected]))
        standard_error_pp = float(np.sqrt(np.sum(
            normalized_weight ** 2 * variance[selected] / sample_used[selected]
        )) * 100)
        low, high = estimate - Z_95 * standard_error_pp, estimate + Z_95 * standard_error_pp
        if limits is not None:
            low, high = max(limits[0], low), min(limits[1], high)
        records.append({
            "data": pd.Timestamp(np.datetime64(int(reference), "D")),
            "estimativa_pct": estimate,
            "ic_baixo_pct": low,
            "ic_alto_pct": high,
            "se_pp": standard_error_pp,
        })
    return SmoothedSeries(pd.DataFrame.from_records(records).set_index("data"), raw)


def _plot_government(composition: dict[str, SmoothedSeries], path: Path) -> None:
    colors = {"Ótimo/Bom": "green", "Regular": "orange", "Ruim/Péssimo": "red"}
    fig, axis = plt.subplots(figsize=(16, 8))
    for label, series in composition.items():
        if not series.raw.empty:
            axis.scatter(series.raw["data_referencia"], series.raw["valor_pct"], color=colors[label], alpha=0.15, s=20)
        if series.values.empty:
            continue
        axis.plot(series.values.index, series.values["estimativa_pct"], color=colors[label], linewidth=2, label=label)
        axis.fill_between(
            series.values.index, series.values["ic_baixo_pct"], series.values["ic_alto_pct"],
            color=colors[label], alpha=0.15,
        )
        latest = series.values.iloc[-1]
        axis.text(series.values.index[-1], latest["estimativa_pct"], f" {latest['estimativa_pct']:.2f}%", color=colors[label], va="center")
    axis.set_title("Avaliação do Governo Lula III", fontsize=16)
    axis.set_xlabel("Data")
    axis.set_ylabel("Percentual")
    axis.grid(True, alpha=0.35)
    axis.legend()
    axis.set_xlim(pd.Timestamp("2023-02-01"), None)
    fig.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def _plot_net(net_15: SmoothedSeries, net_30: SmoothedSeries, path: Path) -> None:
    fig, axis = plt.subplots(figsize=(14, 7))
    for series in (net_15, net_30):
        if not series.raw.empty:
            axis.scatter(series.raw["data_referencia"], series.raw["valor_pct"], color="gray", alpha=0.25, s=20)
            break
    for label, color, series in [("Janela 15 dias", "blue", net_15), ("Janela 30 dias", "red", net_30)]:
        if series.values.empty:
            continue
        axis.plot(series.values.index, series.values["estimativa_pct"], color=color, linewidth=2, label=label)
        axis.fill_between(
            series.values.index, series.values["ic_baixo_pct"], series.values["ic_alto_pct"], color=color, alpha=0.15
        )
        latest = series.values.iloc[-1]
        axis.text(series.values.index[-1], latest["estimativa_pct"], f" {latest['estimativa_pct']:.2f}", color=color, va="center")
    axis.axhline(0, color="black", linewidth=0.8)
    axis.set_title("Avaliação Líquida do Governo Lula III (duas séries com IC 95%)", fontsize=16)
    axis.set_xlabel("Data")
    axis.set_ylabel("Pontos percentuais")
    axis.grid(True, alpha=0.35)
    axis.legend()
    axis.set_xlim(pd.Timestamp("2023-02-01"), None)
    fig.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def _plot_second_round(lula: SmoothedSeries, flavio: SmoothedSeries, path: Path) -> None:
    fig, axis = plt.subplots(figsize=(14, 7))
    for label, color, series in [("Lula", "red", lula), ("Flavio", "blue", flavio)]:
        if not series.raw.empty:
            axis.scatter(series.raw["data_referencia"], series.raw["valor_pct"], color=color, alpha=0.25, s=20)
        if series.values.empty:
            continue
        axis.plot(series.values.index, series.values["estimativa_pct"], color=color, linewidth=2, label=label)
        axis.fill_between(
            series.values.index, series.values["ic_baixo_pct"], series.values["ic_alto_pct"], color=color, alpha=0.15
        )
        latest = series.values.iloc[-1]
        axis.text(series.values.index[-1], latest["estimativa_pct"], f" {latest['estimativa_pct']:.2f}%", color=color, va="center")
    axis.set_title("Segundo Turno 2026: Lula vs Flavio (suavização geométrica + IC 95%)", fontsize=16)
    axis.set_xlabel("Data")
    axis.set_ylabel("Percentual")
    axis.set_ylim(40, 60)
    axis.grid(True, alpha=0.35)
    axis.legend()
    fig.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def _write_workbook(path: Path, sheets: dict[str, pd.DataFrame], index: bool = True) -> None:
    """Escreve as tabelas no formato simples das planilhas legadas."""
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd", date_format="yyyy-mm-dd") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=index)
    # Garante que datas continuem sendo datas reais e sejam exibidas de modo uniforme.
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (pd.Timestamp,)):
                    cell.number_format = "yyyy-mm-dd"
                elif hasattr(cell.value, "year") and hasattr(cell.value, "month") and hasattr(cell.value, "day"):
                    cell.number_format = "yyyy-mm-dd"
    workbook.save(path)


def _composition_tables(composition: dict[str, SmoothedSeries]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    smooth, low, high = [], [], []
    for label, series in composition.items():
        smooth.append(series.values["estimativa_pct"].rename(label))
        low.append(series.values["ic_baixo_pct"].rename(label))
        high.append(series.values["ic_alto_pct"].rename(label))
    result = tuple(pd.concat(parts, axis=1).sort_index() for parts in (smooth, low, high))
    for frame in result:
        frame.index.name = "data"
    return result  # type: ignore[return-value]


def _second_round_rows(contests: pd.DataFrame) -> pd.DataFrame:
    national = _total_brasil(contests)
    mask = (
        national["turno"].map(normalized_text).str.startswith("2", na=False)
        & national["tipo_cenario"].map(normalized_text).eq("estimulada")
        & national["base_voto"].map(normalized_text).eq("totais")
        & national["adversario"].map(normalized_text).str.contains("flavio", na=False)
    )
    rows = national.loc[mask].copy()
    if rows.empty:
        raise ConfigurationError(
            "A base não possui cenário nacional total, estimulado, de 2º turno entre Lula e Flávio. "
            "O Modelo A legado não foi substituído."
        )
    return rows


def _weekly_tracking_from_daily(daily_tracking: pd.DataFrame) -> pd.DataFrame:
    """Reconstrói a média móvel semanal que o arquivo legado chamava de mm7d.

    O legado não atribuía um peso completo a cada leitura diária do tracking.
    Ele usava uma única observação por semana, calculada com os sete dias
    completos e datada no domingo. A agregação é feita separadamente por
    instituto e pelas demais dimensões do cenário para não misturar séries.
    """
    if daily_tracking.empty:
        return daily_tracking.copy()
    group_columns = [
        "instituto", "turno", "tipo_cenario", "cenario", "scenario_id", "adversario",
        "nivel_geografico", "geografia", "segmento_tipo", "segmento", "base_voto", "tipo_pesquisa",
    ]
    value_columns = [
        "lula_pct", "adversario_pct", "outros_candidatos_pct", "brancos_nulos_indecisos_pct",
        "amostra_total", "amostra_segmento", "peso_legacy",
    ]
    records: list[dict[object, object]] = []
    for _, curve in daily_tracking.groupby(group_columns, dropna=False, sort=False):
        curve = curve.sort_values("data_referencia").copy()
        numeric = curve.set_index("data_referencia")[value_columns].apply(pd.to_numeric, errors="coerce")
        weekly = numeric.resample("W-SUN").mean()
        complete_days = curve.set_index("data_referencia").resample("W-SUN").size()
        weekly = weekly.loc[complete_days.reindex(weekly.index).eq(7)]
        if weekly.empty:
            continue
        template = curve.iloc[0].to_dict()
        for date, averages in weekly.iterrows():
            record = template.copy()
            record["data_referencia"] = pd.Timestamp(date)
            record["frequencia_original"] = "semanal"
            for column, value in averages.items():
                record[column] = value
            records.append(record)
    if not records:
        return daily_tracking.iloc[0:0].copy()
    return pd.DataFrame.from_records(records, columns=daily_tracking.columns)


def _legacy_second_round_input(rows: pd.DataFrame) -> pd.DataFrame:
    """Aplica a composição de pesquisas do gráfico histórico Lula × Flávio."""
    regular = rows.loc[rows["tipo_pesquisa"].map(normalized_text).eq("regular")].copy()
    tracking = rows.loc[rows["tipo_pesquisa"].map(normalized_text).eq("tracking")].copy()
    daily_tracking = tracking.loc[tracking["frequencia_original"].map(normalized_text).eq("diaria")].copy()
    non_daily_tracking = tracking.loc[~tracking["frequencia_original"].map(normalized_text).eq("diaria")].copy()
    weekly_tracking = _weekly_tracking_from_daily(daily_tracking)
    result = pd.concat([regular, weekly_tracking, non_daily_tracking], ignore_index=True)
    return result.sort_values("data_referencia").reset_index(drop=True)


def _valid_votes(rows: pd.DataFrame) -> pd.DataFrame:
    """Converte o confronto de dois candidatos para votos válidos, sem residual."""
    result = rows.copy()
    lula = pd.to_numeric(result["lula_pct"], errors="coerce")
    opponent = pd.to_numeric(result["adversario_pct"], errors="coerce")
    denominator = lula + opponent
    result = result.loc[denominator.gt(0)].copy()
    denominator = denominator.loc[result.index]
    result["lula_validos_pct"] = lula.loc[result.index] / denominator * 100
    result["adversario_validos_pct"] = opponent.loc[result.index] / denominator * 100
    return result


def _required(series: SmoothedSeries, label: str) -> SmoothedSeries:
    if series.values.empty:
        raise ConfigurationError(f"Não há observações válidas para gerar '{label}'.")
    return series


def generate_legacy_model_a(config: Config, output_dir: Path, data: StandardizedData) -> list[Path]:
    """Gera apenas as três planilhas e os três gráficos macro do legado."""
    output_dir.mkdir(parents=True, exist_ok=True)
    government = _total_brasil(data.government)
    composition = {
        "Ótimo/Bom": _required(_smooth(government, "otimo_bom_pct", config, 30), "Ótimo/Bom"),
        "Regular": _required(_smooth(government, "regular_pct", config, 30), "Regular"),
        "Ruim/Péssimo": _required(_smooth(government, "ruim_pessimo_pct", config, 30), "Ruim/Péssimo"),
    }
    net_source = government.loc[
        government["otimo_bom_pct"].map(is_present) & government["ruim_pessimo_pct"].map(is_present)
    ].copy()
    net_source["avaliacao_liquida_pct"] = (
        pd.to_numeric(net_source["otimo_bom_pct"], errors="coerce")
        - pd.to_numeric(net_source["ruim_pessimo_pct"], errors="coerce")
    )
    net_15 = _required(
        _smooth(net_source, "avaliacao_liquida_pct", config, 15, ("otimo_bom_pct", "ruim_pessimo_pct"), (-100.0, 100.0)),
        "Avaliação líquida em 15 dias",
    )
    net_30 = _required(
        _smooth(net_source, "avaliacao_liquida_pct", config, 30, ("otimo_bom_pct", "ruim_pessimo_pct"), (-100.0, 100.0)),
        "Avaliação líquida em 30 dias",
    )
    second_round = _valid_votes(_legacy_second_round_input(_second_round_rows(data.contests)))
    lula = _required(_smooth(second_round, "lula_validos_pct", config, 30), "Lula no 2º turno")
    flavio = _required(_smooth(second_round, "adversario_validos_pct", config, 30), "Flávio no 2º turno")

    composition_smooth, composition_low, composition_high = _composition_tables(composition)
    government_workbook = output_dir / f"{GOVERNMENT_OUTPUT}.xlsx"
    _write_workbook(government_workbook, {
        "Suavizados": composition_smooth,
        "IC_Baixo": composition_low,
        "IC_Alto": composition_high,
    })

    net_tables: dict[str, pd.DataFrame] = {}
    for sheet_name, series in [("Serie_Janela_15d", net_15), ("Serie_Janela_30d", net_30)]:
        table = series.values.reset_index().rename(columns={
            "estimativa_pct": "aprov_liq_suavizado", "ic_baixo_pct": "ci_low", "ic_alto_pct": "ci_high",
        })
        net_tables[sheet_name] = table[["data", "aprov_liq_suavizado", "ci_low", "ci_high", "se_pp"]]
    net_workbook = output_dir / f"{NET_GOVERNMENT_OUTPUT}.xlsx"
    _write_workbook(net_workbook, net_tables, index=False)

    second_round_tables: dict[str, pd.DataFrame] = {}
    for sheet_name, column in [("Suavizados", "estimativa_pct"), ("IC_Baixo", "ic_baixo_pct"), ("IC_Alto", "ic_alto_pct")]:
        table = pd.concat([lula.values[column].rename("Lula"), flavio.values[column].rename("Flavio")], axis=1).sort_index()
        table.index.name = "data"
        second_round_tables[sheet_name] = table
    second_round_workbook = output_dir / f"{SECOND_ROUND_OUTPUT}.xlsx"
    _write_workbook(second_round_workbook, second_round_tables)

    government_chart = output_dir / f"{GOVERNMENT_OUTPUT}.png"
    net_chart = output_dir / f"{NET_GOVERNMENT_OUTPUT}.png"
    second_round_chart = output_dir / f"{SECOND_ROUND_OUTPUT}.png"
    _plot_government(composition, government_chart)
    _plot_net(net_15, net_30, net_chart)
    _plot_second_round(lula, flavio, second_round_chart)
    return [government_workbook, net_workbook, second_round_workbook, government_chart, net_chart, second_round_chart]
