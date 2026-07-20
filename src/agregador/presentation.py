from __future__ import annotations

from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter
import numpy as np
import pandas as pd

from .configuration import Config
from .domain import safe_slug
from .standardization import StandardizedData


BLUE = "#155e75"
BLUE_LIGHT = "#67a9cf"
RED = "#b91c1c"
ORANGE = "#c2410c"
GRAY = "#64748b"
GREEN = "#15803d"


def _empty(columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=columns)


def _date_label(value: Any) -> str:
    if pd.isna(value):
        return "sem data"
    return pd.Timestamp(value).strftime("%Y-%m-%d")


def _pct(value: Any) -> str:
    if pd.isna(value):
        return "n/d"
    return f"{float(value):.1f}%".replace(".", ",")


def _pp(value: Any) -> str:
    if pd.isna(value):
        return "n/d"
    sign = "+" if float(value) > 0 else ""
    return f"{sign}{float(value):.1f} p.p.".replace(".", ",")


def _filter_brazil_total(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    result = frame.copy()
    return result.loc[
        result["nivel_geografico"].eq("Brasil")
        & result["geografia"].eq("Brasil")
        & result["segmento_tipo"].eq("Total")
        & result["segmento"].eq("Total")
    ].copy()


def _second_turn_base(aggregates: pd.DataFrame) -> pd.DataFrame:
    if aggregates.empty:
        return aggregates.copy()
    result = _filter_brazil_total(aggregates)
    return result.loc[
        result["product"].eq("Confrontos")
        & result["turno"].astype(str).str.startswith("2")
        & result["vote_base"].eq("Totais")
        & result["tipo_cenario"].eq("Estimulada")
    ].copy()


def _wide_second_turn(aggregates: pd.DataFrame) -> pd.DataFrame:
    base = _second_turn_base(aggregates)
    if base.empty:
        return _empty(["adversario", "scenario_id", "reference_date"])
    metrics = ["lula_totais", "adversario_totais", "nao_voto_totais", "margem_totais"]
    base = base.loc[base["metric"].isin(metrics)].copy()
    keys = ["adversario", "scenario_id", "reference_date", "poll_count", "institute_count"]
    if "evidence_status" in base.columns:
        keys.append("evidence_status")
    wide = base.pivot_table(index=keys, columns="metric", values="estimate_pct", aggfunc="first").reset_index()
    wide.columns.name = None
    for column in metrics:
        if column not in wide:
            wide[column] = np.nan
    # O Modelo B publica incerteza e status de evidência por ponto. Mantemos
    # esses campos ao lado da margem para que o gráfico executivo possa
    # mostrar a faixa experimental sem alterar a série do Modelo A.
    extra_columns = [
        "standard_error_pp", "interval_low_pct", "interval_high_pct",
        "effective_information", "concentration_cap_feasible",
    ]
    available_extra = [column for column in extra_columns if column in base.columns]
    if available_extra and "margem_totais" in base["metric"].unique():
        margin_extra = base.loc[base["metric"].eq("margem_totais"), keys + available_extra].drop_duplicates(keys)
        wide = wide.merge(margin_extra, on=keys, how="left")
    return wide.sort_values(["adversario", "reference_date"])


def _latest_rows(frame: pd.DataFrame, group_columns: list[str]) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    result = frame.sort_values("reference_date").copy()
    return result.loc[result.groupby(group_columns, dropna=False)["reference_date"].transform("max").eq(result["reference_date"])].copy()


def _break_long_gaps(curve: pd.DataFrame, value_columns: list[str], maximum_gap_days: int = 45) -> pd.DataFrame:
    result = curve.sort_values("reference_date").copy()
    dates = pd.to_datetime(result["reference_date"])
    gaps = dates.diff().dt.days.gt(maximum_gap_days).fillna(False)
    for column in value_columns:
        result.loc[gaps, column] = np.nan
    return result


def _friendly_opponent(value: Any) -> str:
    return str(value) if pd.notna(value) and str(value).strip() else "adversário não informado"


def _save_figure(figure: plt.Figure, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(figure)


def _plot_margin_curves(wide: pd.DataFrame, directory: Path) -> list[Path]:
    outputs: list[Path] = []
    if wide.empty or "margem_totais" not in wide:
        return outputs
    for opponent, curve in wide.groupby("adversario", dropna=False, sort=True):
        curve = _break_long_gaps(curve, ["margem_totais"])
        if curve["margem_totais"].notna().sum() == 0:
            continue
        name = _friendly_opponent(opponent)
        latest = curve.dropna(subset=["margem_totais"]).iloc[-1]
        figure, axis = plt.subplots(figsize=(11, 5.6))
        axis.plot(curve["reference_date"], curve["margem_totais"], color=BLUE, linewidth=2.4)
        if {"interval_low_pct", "interval_high_pct"}.issubset(curve.columns):
            lower = pd.to_numeric(curve["interval_low_pct"], errors="coerce")
            upper = pd.to_numeric(curve["interval_high_pct"], errors="coerce")
            if lower.notna().any() and upper.notna().any():
                axis.fill_between(
                    mdates.date2num(pd.to_datetime(curve["reference_date"])), lower, upper,
                    color=BLUE_LIGHT, alpha=0.24, label="intervalo experimental",
                )
        axis.axhline(0, color=GRAY, linewidth=1)
        axis.scatter([latest["reference_date"]], [latest["margem_totais"]], color=ORANGE, s=45, zorder=3)
        axis.annotate(
            f"Último: {_pp(latest['margem_totais'])}\n{_date_label(latest['reference_date'])}",
            (latest["reference_date"], latest["margem_totais"]), textcoords="offset points", xytext=(8, 8),
            fontsize=10, color=ORANGE,
        )
        axis.set_title(f"Vantagem de Lula sobre {name}")
        axis.set_ylabel("pontos percentuais — votos totais")
        axis.set_xlabel("Data de referência")
        axis.grid(axis="y", alpha=0.22)
        if "interval_low_pct" in curve.columns and curve["interval_low_pct"].notna().any():
            axis.legend(loc="best")
        axis.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{value:.0f}"))
        output = directory / f"lula_vs_{safe_slug(name)}" / "margem_votos_totais.png"
        _save_figure(figure, output)
        outputs.append(output)
    return outputs


def _plot_support_and_nonvote(wide: pd.DataFrame, directory: Path) -> list[Path]:
    outputs: list[Path] = []
    if wide.empty:
        return outputs
    for opponent, curve in wide.groupby("adversario", dropna=False, sort=True):
        columns = [column for column in ["lula_totais", "adversario_totais", "nao_voto_totais"] if column in curve]
        if not columns or curve[columns].notna().sum().sum() == 0:
            continue
        curve = _break_long_gaps(curve, columns)
        name = _friendly_opponent(opponent)
        figure, axis = plt.subplots(figsize=(11, 5.6))
        labels = {"lula_totais": "Lula", "adversario_totais": name, "nao_voto_totais": "Brancos, nulos e indecisos"}
        colors = {"lula_totais": BLUE, "adversario_totais": RED, "nao_voto_totais": GRAY}
        styles = {"lula_totais": "-", "adversario_totais": "-", "nao_voto_totais": "--"}
        for column in columns:
            axis.plot(curve["reference_date"], curve[column], label=labels[column], color=colors[column], linestyle=styles[column], linewidth=2)
        axis.set_ylim(0, 100)
        axis.set_title(f"Intenção de voto e não voto — Lula × {name}")
        axis.set_ylabel("percentual publicado/agregado")
        axis.set_xlabel("Data de referência")
        axis.legend(loc="best", frameon=True)
        axis.grid(axis="y", alpha=0.22)
        output = directory / f"lula_vs_{safe_slug(name)}" / "intencao_e_nao_voto.png"
        _save_figure(figure, output)
        outputs.append(output)
    return outputs


def _plot_comparison(wide: pd.DataFrame, directory: Path) -> Path | None:
    if wide.empty or "margem_totais" not in wide:
        return None
    latest = _latest_rows(wide.dropna(subset=["margem_totais"]), ["adversario", "scenario_id"])
    if latest.empty:
        return None
    latest = latest.sort_values("margem_totais")
    figure, axis = plt.subplots(figsize=(10.5, max(4.5, 0.55 * len(latest))))
    colors = [BLUE if value >= 0 else RED for value in latest["margem_totais"]]
    axis.barh(latest["adversario"].map(_friendly_opponent), latest["margem_totais"], color=colors, alpha=0.88)
    axis.axvline(0, color=GRAY, linewidth=1)
    minimum = float(latest["margem_totais"].min())
    maximum = float(latest["margem_totais"].max())
    span = max(1.0, maximum - minimum)
    axis.set_xlim(minimum - max(2.0, 0.08 * span), maximum + max(7.0, 0.28 * span))
    for position, (_, row) in enumerate(latest.iterrows()):
        axis.text(float(row["margem_totais"]) + (0.2 if row["margem_totais"] >= 0 else -0.2), position,
                  f"{_pp(row['margem_totais'])} | {_date_label(row['reference_date'])} | {int(row['institute_count'])} instit.",
                  va="center", ha="left" if row["margem_totais"] >= 0 else "right", fontsize=9)
    axis.set_title("Comparação dos confrontos disponíveis — último ponto")
    axis.set_xlabel("vantagem de Lula, em pontos percentuais")
    axis.grid(axis="x", alpha=0.2)
    output = directory / "painel_comparativo_confrontos.png"
    _save_figure(figure, output)
    return output


def _plot_institute_evidence(
    contests: pd.DataFrame,
    wide: pd.DataFrame,
    directory: Path,
    model_name: str = "Modelo A",
) -> list[Path]:
    outputs: list[Path] = []
    raw = _filter_brazil_total(contests)
    if raw.empty:
        return outputs
    raw = raw.loc[raw["turno"].astype(str).str.startswith("2") & raw["tipo_cenario"].eq("Estimulada")].copy()
    raw["margem_publicada"] = raw["lula_pct"] - raw["adversario_pct"]
    for opponent, curve in raw.groupby("adversario", dropna=False, sort=True):
        curve = curve.dropna(subset=["margem_publicada"])
        if curve.empty:
            continue
        name = _friendly_opponent(opponent)
        figure, axis = plt.subplots(figsize=(11, 5.6))
        institutes = curve["instituto"].value_counts().head(10).index.tolist()
        palette = plt.get_cmap("tab10")
        for index, institute in enumerate(institutes):
            points = curve.loc[curve["instituto"].eq(institute)]
            color = palette(index % 10)
            for poll_type, marker, alpha in [("regular", "o", 0.62), ("tracking", "x", 0.85)]:
                typed = points.loc[points["tipo_pesquisa"].fillna("regular").eq(poll_type)]
                if typed.empty:
                    continue
                label = f"{institute} ({poll_type})"
                axis.scatter(
                    typed["data_referencia"], typed["margem_publicada"], s=34, alpha=alpha,
                    color=color, marker=marker, label=label,
                )
        model = wide.loc[wide["adversario"].eq(opponent)].sort_values("reference_date")
        if not model.empty and "margem_totais" in model:
            axis.plot(model["reference_date"], model["margem_totais"], color="#0f172a", linewidth=2.4, label=model_name)
        axis.axhline(0, color=GRAY, linewidth=1)
        axis.set_title(f"Evidência por instituto — Lula × {name}")
        axis.set_ylabel("margem publicada: Lula − adversário")
        axis.set_xlabel("Data de referência")
        axis.grid(axis="y", alpha=0.2)
        axis.legend(loc="best", fontsize=8)
        output = directory / f"lula_vs_{safe_slug(name)}" / "evidencia_por_instituto.png"
        _save_figure(figure, output)
        outputs.append(output)
    return outputs


def _latest_demographic_profiles(contests: pd.DataFrame) -> pd.DataFrame:
    raw = contests.copy()
    raw = raw.loc[
        raw["turno"].astype(str).str.startswith("2")
        & raw["tipo_cenario"].eq("Estimulada")
        & raw["nivel_geografico"].eq("Brasil")
        & raw["geografia"].eq("Brasil")
        & raw["segmento_tipo"].isin(["Sexo", "Idade", "Religiao"])
    ].copy()
    if raw.empty:
        return raw
    rows: list[pd.DataFrame] = []
    for (opponent, segment_type), subset in raw.groupby(["adversario", "segmento_tipo"], dropna=False):
        candidates = subset.groupby(["poll_group_id", "data_referencia", "instituto"], dropna=False).agg(segment_count=("segmento", "nunique")).reset_index()
        candidates = candidates.sort_values(["data_referencia", "segment_count"], ascending=[False, False])
        if candidates.empty:
            continue
        chosen = candidates.iloc[0]
        chosen_rows = subset.loc[
            subset["poll_group_id"].eq(chosen["poll_group_id"])
            & subset["data_referencia"].eq(chosen["data_referencia"])
            & subset["instituto"].eq(chosen["instituto"])
        ].copy()
        chosen_rows["opponent"] = opponent
        chosen_rows["segment_type"] = segment_type
        chosen_rows["margin_published"] = chosen_rows["lula_pct"] - chosen_rows["adversario_pct"]
        chosen_rows["profile_status"] = "fotografia_da_onda_publicada"
        rows.append(chosen_rows)
    return pd.concat(rows, ignore_index=True) if rows else raw.iloc[0:0].copy()


def _plot_demographics(profiles: pd.DataFrame, directory: Path) -> list[Path]:
    outputs: list[Path] = []
    if profiles.empty:
        return outputs
    for (opponent, segment_type), frame in profiles.groupby(["opponent", "segment_type"], dropna=False, sort=True):
        frame = frame.sort_values("margin_published")
        name = _friendly_opponent(opponent)
        figure, axis = plt.subplots(figsize=(9.5, max(4, 0.45 * len(frame))))
        colors = [BLUE if value >= 0 else RED for value in frame["margin_published"]]
        axis.barh(frame["segmento"], frame["margin_published"], color=colors, alpha=0.88)
        axis.axvline(0, color=GRAY, linewidth=1)
        minimum = float(frame["margin_published"].min())
        maximum = float(frame["margin_published"].max())
        span = max(1.0, maximum - minimum)
        axis.set_xlim(minimum - max(3.0, 0.18 * span), maximum + max(3.0, 0.18 * span))
        for position, (_, row) in enumerate(frame.iterrows()):
            axis.text(float(row["margin_published"]) + (0.2 if row["margin_published"] >= 0 else -0.2), position,
                      _pp(row["margin_published"]), va="center", ha="left" if row["margin_published"] >= 0 else "right", fontsize=9)
        axis.set_title(f"Perfil publicado por {segment_type.lower()} — Lula × {name}")
        axis.set_xlabel("margem Lula − adversário (pontos percentuais)")
        axis.grid(axis="x", alpha=0.2)
        output = directory / f"lula_vs_{safe_slug(name)}" / f"perfil_{safe_slug(segment_type)}.png"
        _save_figure(figure, output)
        outputs.append(output)
    return outputs


def _demographic_export(profiles: pd.DataFrame) -> pd.DataFrame:
    """Mantém na aba de apresentação somente os campos úteis ao leitor."""
    columns = [
        "opponent", "segment_type", "segmento", "data_referencia", "instituto",
        "tipo_pesquisa", "frequencia_original", "lula_pct", "adversario_pct",
        "margin_published", "brancos_nulos_indecisos_pct", "amostra_total",
        "amostra_segmento", "profile_status",
    ]
    available = [column for column in columns if column in profiles.columns]
    return profiles.loc[:, available].copy()


def _first_turn_scenarios(contests: pd.DataFrame) -> pd.DataFrame:
    raw = contests.loc[
        contests["turno"].astype(str).str.startswith("1")
        & contests["nivel_geografico"].eq("Brasil")
        & contests["geografia"].eq("Brasil")
        & contests["segmento_tipo"].eq("Total")
        & contests["segmento"].eq("Total")
    ].copy()
    if raw.empty:
        return _empty(["data_referencia", "instituto", "tipo_cenario", "cenario", "candidatos", "n_candidatos", "reconstrucao_pct", "status"])
    records: list[dict[str, Any]] = []
    group_columns = ["poll_group_id", "scenario_id", "data_referencia", "instituto", "tipo_cenario", "cenario", "scenario_composition_id"]
    for keys, frame in raw.groupby(group_columns, dropna=False, sort=True):
        payload = dict(zip(group_columns, keys, strict=True))
        candidates = sorted(frame["adversario"].dropna().astype(str).unique().tolist())
        lula = float(frame["lula_pct"].dropna().iloc[0]) if frame["lula_pct"].notna().any() else np.nan
        opponent_sum = float(frame["adversario_pct"].sum(min_count=1)) if frame["adversario_pct"].notna().any() else np.nan
        nonvote = float(frame["brancos_nulos_indecisos_pct"].dropna().iloc[0]) if frame["brancos_nulos_indecisos_pct"].notna().any() else np.nan
        reconstructed = lula + opponent_sum + nonvote if pd.notna(lula) and pd.notna(opponent_sum) and pd.notna(nonvote) else np.nan
        payload.update({
            "candidatos": " | ".join(candidates), "n_candidatos": len(candidates),
            "lula_pct": lula, "adversarios_somados_pct": opponent_sum,
            "brancos_nulos_indecisos_pct": nonvote, "reconstrucao_pct": reconstructed,
            "status": "composicao_reconciliada" if pd.notna(reconstructed) and abs(reconstructed - 100) <= 0.15 else "nao_comparar_como_composicao_completa",
        })
        records.append(payload)
    return pd.DataFrame.from_records(records).sort_values("data_referencia")


def _plot_first_turn_matrix(scenarios: pd.DataFrame, directory: Path) -> Path | None:
    if scenarios.empty:
        return None
    candidates = sorted({candidate for value in scenarios["candidatos"].fillna("") for candidate in str(value).split(" | ") if candidate})
    if not candidates:
        return None
    matrix = np.zeros((len(scenarios), len(candidates)))
    for row_index, value in enumerate(scenarios["candidatos"].fillna("")):
        present = set(str(value).split(" | "))
        for col_index, candidate in enumerate(candidates):
            matrix[row_index, col_index] = 1 if candidate in present else np.nan
    labels = [f"{_date_label(row.data_referencia)} — {row.instituto} — {row.cenario}" for row in scenarios.itertuples()]
    figure, axis = plt.subplots(figsize=(max(10, 0.55 * len(candidates)), max(4.5, 0.35 * len(scenarios))))
    axis.imshow(matrix, aspect="auto", cmap="Blues", vmin=0, vmax=1)
    axis.set_xticks(range(len(candidates)), candidates, rotation=45, ha="right")
    axis.set_yticks(range(len(labels)), labels)
    axis.set_title("Primeiro turno — quais candidatos aparecem em cada cenário")
    axis.set_xlabel("candidato explicitamente publicado")
    axis.set_ylabel("data, instituto e cenário")
    figure.tight_layout()
    output = directory / "primeiro_turno_mapa_de_cenarios.png"
    _save_figure(figure, output)
    return output


def _plot_government(aggregates: pd.DataFrame, directory: Path) -> list[Path]:
    gov = _filter_brazil_total(aggregates)
    gov = gov.loc[gov["product"].eq("Lula3")].copy()
    if gov.empty:
        return []
    outputs: list[Path] = []
    labels = {"otimo_bom": "Ótimo/bom", "regular": "Regular", "ruim_pessimo": "Ruim/péssimo"}
    colors = {"otimo_bom": GREEN, "regular": ORANGE, "ruim_pessimo": RED}
    composition = gov.loc[gov["metric"].isin(labels)].pivot_table(index="reference_date", columns="metric", values="estimate_pct", aggfunc="first").reset_index()
    figure, axis = plt.subplots(figsize=(11, 5.6))
    for metric, label in labels.items():
        if metric in composition:
            axis.plot(composition["reference_date"], composition[metric], label=label, color=colors[metric], linewidth=2)
    axis.set_ylim(0, 100)
    axis.set_title("Avaliação do governo — composição da avaliação")
    axis.set_ylabel("percentual")
    axis.set_xlabel("Data de referência")
    axis.legend(loc="best")
    axis.grid(axis="y", alpha=0.22)
    output = directory / "governo_composicao_avaliacao.png"
    _save_figure(figure, output)
    outputs.append(output)
    net = gov.loc[gov["metric"].eq("favorabilidade_liquida")].sort_values("reference_date")
    if not net.empty:
        figure, axis = plt.subplots(figsize=(11, 5.6))
        axis.plot(net["reference_date"], net["estimate_pct"], color=BLUE, linewidth=2.4)
        axis.axhline(0, color=GRAY, linewidth=1)
        axis.set_title("Favorabilidade líquida do governo")
        axis.set_ylabel("ótimo/bom − ruim/péssimo (p.p.)")
        axis.set_xlabel("Data de referência")
        axis.grid(axis="y", alpha=0.22)
        output = directory / "governo_favorabilidade_liquida.png"
        _save_figure(figure, output)
        outputs.append(output)
    return outputs


def _coverage_table(
    wide: pd.DataFrame,
    weights: pd.DataFrame,
    min_polls: int = 3,
    min_institutes: int = 2,
) -> pd.DataFrame:
    if wide.empty:
        return _empty(["adversario", "data_referencia", "pesquisas", "institutos", "tracking_pct", "status"])
    latest = _latest_rows(wide.dropna(subset=["margem_totais"]), ["adversario", "scenario_id"]).copy()
    rows: list[dict[str, Any]] = []
    for _, row in latest.iterrows():
        source = weights.loc[
            weights["product"].eq("Confrontos")
            & weights["metric"].eq("margem_totais")
            & weights["adversario"].eq(row["adversario"])
            & weights["reference_date"].eq(row["reference_date"])
        ].copy()
        polls = source["poll_group_id"].nunique() if not source.empty else int(row.get("poll_count", 0))
        institutes = source["instituto"].nunique() if not source.empty else int(row.get("institute_count", 0))
        tracking = source.drop_duplicates("poll_group_id")["tipo_pesquisa"].eq("tracking").mean() if not source.empty else np.nan
        rows.append({
            "adversario": row["adversario"], "data_referencia": row["reference_date"],
            "estimativa_margem_pp": row["margem_totais"], "pesquisas": int(polls), "institutos": int(institutes),
            "tracking_pct": tracking * 100 if pd.notna(tracking) else np.nan,
            "status": "cobertura suficiente" if polls >= min_polls and institutes >= min_institutes else "cobertura baixa — manter aviso",
        })
    return pd.DataFrame.from_records(rows)


def _model_b_uncertainty_table(aggregates: pd.DataFrame) -> pd.DataFrame:
    """Recorte executivo da incerteza e dos gates próprios do Modelo B."""
    base = _filter_brazil_total(aggregates)
    if base.empty:
        return _empty([
            "adversario", "data_referencia", "margem_totais", "erro_padrao_pp",
            "intervalo_baixo_pp", "intervalo_alto_pp", "informacao_efetiva",
            "cap_concentracao_viavel", "evidencia_status", "pesquisas", "institutos",
        ])
    base = base.loc[
        base["product"].eq("Confrontos")
        & base["turno"].astype(str).str.startswith("2")
        & base["vote_base"].eq("Totais")
        & base["tipo_cenario"].eq("Estimulada")
        & base["metric"].eq("margem_totais")
    ].copy()
    if base.empty:
        return _empty([
            "adversario", "data_referencia", "margem_totais", "erro_padrao_pp",
            "intervalo_baixo_pp", "intervalo_alto_pp", "informacao_efetiva",
            "cap_concentracao_viavel", "evidencia_status", "pesquisas", "institutos",
        ])
    columns = {
        "adversario": "adversario", "reference_date": "data_referencia",
        "estimate_pct": "margem_totais", "standard_error_pp": "erro_padrao_pp",
        "interval_low_pct": "intervalo_baixo_pp", "interval_high_pct": "intervalo_alto_pp",
        "effective_information": "informacao_efetiva",
        "concentration_cap_feasible": "cap_concentracao_viavel",
        "evidence_status": "evidencia_status", "poll_count": "pesquisas",
        "institute_count": "institutos",
    }
    available = [column for column in columns if column in base.columns]
    result = base.loc[:, available].rename(columns={column: columns[column] for column in available})
    return result.sort_values(["adversario", "data_referencia"])


def _probability_export(probabilities: pd.DataFrame) -> pd.DataFrame:
    if probabilities is None or probabilities.empty:
        return _empty([
            "adversario", "data_referencia", "probabilidade_lula", "experimental",
            "base_voto", "motivo",
        ])
    frame = probabilities.copy()
    rename = {
        "opponent": "adversario", "reference_date": "data_referencia",
        "probability_lula": "probabilidade_lula", "experimental_flag": "experimental",
        "vote_base": "base_voto", "reason": "motivo",
    }
    frame = frame.rename(columns={column: rename[column] for column in rename if column in frame.columns})
    wanted = [
        "adversario", "data_referencia", "probabilidade_lula", "experimental",
        "base_voto", "motivo", "probability_type", "geographic_level", "geography",
        "segment_type", "segment", "scenario_id",
    ]
    return frame.loc[:, [column for column in wanted if column in frame.columns]].sort_values(
        [column for column in ["adversario", "data_referencia"] if column in frame.columns]
    )


def _plot_probability_curves(probabilities: pd.DataFrame, directory: Path) -> list[Path]:
    outputs: list[Path] = []
    if probabilities is None or probabilities.empty:
        return outputs
    frame = probabilities.copy()
    if "opponent" not in frame.columns or "probability_lula" not in frame.columns:
        return outputs
    filters = pd.Series(True, index=frame.index)
    for column, value in [("geographic_level", "Brasil"), ("geography", "Brasil"), ("segment_type", "Total"), ("segment", "Total"), ("vote_base", "Totais")]:
        if column in frame.columns:
            filters &= frame[column].eq(value)
    frame = frame.loc[filters & frame["probability_lula"].notna()].copy()
    if frame.empty:
        return outputs
    for opponent, curve in frame.groupby("opponent", dropna=False, sort=True):
        curve = curve.sort_values("reference_date")
        name = _friendly_opponent(opponent)
        latest = curve.iloc[-1]
        figure, axis = plt.subplots(figsize=(11, 5.6))
        axis.plot(curve["reference_date"], curve["probability_lula"] * 100, color=BLUE, linewidth=2.4)
        axis.scatter([latest["reference_date"]], [latest["probability_lula"] * 100], color=ORANGE, s=45, zorder=3)
        axis.annotate(
            f"Última: {_pct(latest['probability_lula'] * 100)}\n{_date_label(latest['reference_date'])}",
            (latest["reference_date"], latest["probability_lula"] * 100), textcoords="offset points", xytext=(8, 8),
            fontsize=10, color=ORANGE,
        )
        axis.set_ylim(0, 100)
        axis.set_title(f"Probabilidade experimental de liderança de Lula — {name}")
        axis.set_ylabel("probabilidade (%)")
        axis.set_xlabel("Data de referência")
        axis.grid(axis="y", alpha=0.22)
        output = directory / f"lula_vs_{safe_slug(name)}" / "probabilidade_lula_experimental.png"
        _save_figure(figure, output)
        outputs.append(output)
    return outputs


def _summary_table(aggregates: pd.DataFrame, coverage: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    gov = _filter_brazil_total(aggregates)
    net = gov.loc[gov["product"].eq("Lula3") & gov["metric"].eq("favorabilidade_liquida")].sort_values("reference_date")
    if not net.empty:
        item = net.iloc[-1]
        rows.append({"bloco": "Contexto do governo", "indicador": "Favorabilidade líquida", "data": item["reference_date"], "estimativa": item["estimate_pct"], "margem": item["estimate_pct"], "pesquisas": item["poll_count"], "institutos": item["institute_count"], "status": "métrica de avaliação; não é intenção de voto"})
    for _, item in coverage.iterrows():
        rows.append({"bloco": "Segundo turno", "indicador": f"Lula × {item['adversario']}", "data": item["data_referencia"], "estimativa": item["estimativa_margem_pp"], "margem": item["estimativa_margem_pp"], "pesquisas": item["pesquisas"], "institutos": item["institutos"], "status": item["status"]})
    return pd.DataFrame.from_records(rows, columns=["bloco", "indicador", "data", "estimativa", "margem", "pesquisas", "institutos", "status"])


def _notes(model_label: str, config: Config) -> pd.DataFrame:
    notes = [
        {"topico": "Modelo", "nota": f"{model_label}; as médias são estimativas do agregador, não previsões eleitorais."},
        {"topico": "Segundo turno", "nota": "Cada adversário é tratado como uma série independente. Datas, cenário, geografia, segmento e base de voto não são misturados."},
        {"topico": "Votos válidos", "nota": "Votos válidos derivados devem aparecer em aba/gráfico separado dos votos totais."},
        {"topico": "Tracking", "nota": "Tracking é identificado na cobertura. Pontos diários não devem ser interpretados como institutos independentes."},
        {"topico": "Aberturas", "nota": "O perfil demográfico é uma fotografia da onda publicada escolhida; não combina rótulos diferentes automaticamente."},
        {"topico": "Primeiro turno", "nota": "Não há média entre listas de candidatos diferentes. O mapa mostra presença dos candidatos e a planilha registra a composição publicada."},
        {"topico": "Fonte", "nota": f"Entrada: {config.value('input_file')}; referência: data_referencia; versão: {model_label}."},
    ]
    if str(config.value("methodology_version", "")).startswith("model-b"):
        notes.insert(1, {"topico": "Modelo B", "nota": "Candidato experimental: meia-vida, teto de concentração por instituto, piso de incerteza e gates de cobertura. Não substituir o Modelo A nem interpretar a probabilidade como previsão eleitoral."})
        notes.insert(4, {"topico": "Probabilidade", "nota": "A aba Probabilidade_Experimental só tem estimativa quando os gates de evidência são atendidos; células sem valor carregam o motivo da supressão. O gráfico usa Brasil/Total/Totais; a aba também preserva outros recortes publicados."})
    return pd.DataFrame(notes)


def _format_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="155E75")
    header_font = Font(color="FFFFFF", bold=True)
    date_columns = {"data", "data_referencia", "reference_date", "inicio", "fim"}
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 28
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells[:100]]
            width = min(42, max(12, max((len(value) for value in values), default=10) + 2))
            sheet.column_dimensions[letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in date_columns or str(sheet.cell(1, cell.column).value) in date_columns:
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, (int, float)) and cell.value is not None:
                    header = str(sheet.cell(1, cell.column).value)
                    cell.number_format = "0.0" if ("pct" in header or "margem" in header or "estimativa" in header or "tracking" in header) else "#,##0"
    workbook.save(path)


def generate_presentation(
    config: Config,
    run_dir: Path,
    data: StandardizedData,
    aggregates: pd.DataFrame,
    weights: pd.DataFrame,
    probabilities: pd.DataFrame | None = None,
    model_label: str = "Modelo A — baseline oficial",
) -> Path:
    """Gera gráficos e planilha de apresentação sem fixar nomes de adversários."""
    root = run_dir / "apresentacao"
    charts = root / "graficos"
    root.mkdir(parents=True, exist_ok=True)
    is_model_b = str(config.value("methodology_version", "")).startswith("model-b")
    wide = _wide_second_turn(aggregates)
    coverage_config = config.nested("model_b").get("coverage", {}) if is_model_b else config.nested("coverage", {})
    coverage = _coverage_table(
        wide,
        weights,
        min_polls=int(coverage_config.get("min_polls", 3)),
        min_institutes=int(coverage_config.get("min_institutes", 2)),
    )
    profiles = _latest_demographic_profiles(data.contests)
    profiles_export = _demographic_export(profiles)
    scenarios = _first_turn_scenarios(data.contests)
    uncertainty = _model_b_uncertainty_table(aggregates) if is_model_b else _empty([])
    probability_export = _probability_export(probabilities) if is_model_b else _empty([])
    generated = []
    generated.extend(_plot_margin_curves(wide, charts))
    generated.extend(_plot_support_and_nonvote(wide, charts))
    comparison = _plot_comparison(wide, charts)
    if comparison:
        generated.append(comparison)
    generated.extend(_plot_institute_evidence(data.contests, wide, charts, model_name="Modelo B" if is_model_b else "Modelo A"))
    generated.extend(_plot_demographics(profiles, charts / "aberturas"))
    matrix = _plot_first_turn_matrix(scenarios, charts)
    if matrix:
        generated.append(matrix)
    generated.extend(_plot_government(aggregates, charts))
    if is_model_b:
        generated.extend(_plot_probability_curves(probabilities, charts))
    workbook_path = root / ("Resultados_Modelo_B.xlsx" if is_model_b else "Resultados_Modelo_A.xlsx")
    summary = _summary_table(aggregates, coverage)
    with pd.ExcelWriter(workbook_path, engine="openpyxl", date_format="YYYY-MM-DD", datetime_format="YYYY-MM-DD") as writer:
        summary.to_excel(writer, sheet_name="Resumo_Executivo", index=False)
        wide.to_excel(writer, sheet_name="Segundo_Turno", index=False)
        profiles_export.to_excel(writer, sheet_name="Aberturas", index=False)
        scenarios.to_excel(writer, sheet_name="Primeiro_Turno", index=False)
        coverage.to_excel(writer, sheet_name="Cobertura", index=False)
        if is_model_b:
            uncertainty.to_excel(writer, sheet_name="Incerteza_Modelo_B", index=False)
            probability_export.to_excel(writer, sheet_name="Probabilidade_Experimental", index=False)
        aggregates.to_excel(writer, sheet_name=("Agregados_Modelo_B" if is_model_b else "Agregados_Modelo_A"), index=False)
        _notes(model_label, config).to_excel(writer, sheet_name="Notas_Metodologicas", index=False)
    _format_workbook(workbook_path)
    manifest = root / "LEIA_ME.txt"
    workbook_name = "Resultados_Modelo_B.xlsx" if is_model_b else "Resultados_Modelo_A.xlsx"
    manifest.write_text(
        "Resultados de apresentação gerados automaticamente.\n\n"
        f"Comece por {workbook_name} > Resumo_Executivo.\n"
        "Os gráficos em graficos/ usam nomes genéricos e são gerados para cada adversário disponível.\n"
        "Não misture votos totais com votos válidos derivados. Consulte Notas_Metodologicas antes de interpretar.\n",
        encoding="utf-8",
    )
    return root
