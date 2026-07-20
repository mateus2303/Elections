from __future__ import annotations

import json
import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "Número do registro", "Eleição", "Data/hora de registro", "Data permitida para divulgação", "UF", "Abrangência", "Município/unidade", "Cargo(s)", "Instituto", "Nome fantasia", "CNPJ", "Contratante(s)", "Pagante(s)", "Valor da pesquisa", "Amostra", "Período de campo", "Metodologia", "Plano amostral, margem e confiança", "Estatístico", "CONRE", "Situação da disponibilidade", "Status de divulgação", "Primeira localização", "Última verificação", "Última alteração", "Publicação confirmada em", "Fonte da publicação", "Observações manuais", "Tags", "Prioridade", "Referência do questionário", "Referência territorial", "Referência da nota fiscal", "Referência oficial", "Hash"
]


def _dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
        return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
    except ValueError:
        return None


def _money(cents: int | None) -> float | None:
    return None if cents is None else cents / 100


def _join_names(items: list[dict[str, Any]], field: str) -> str | None:
    values = [str(x.get(field)) for x in items if x.get(field)]
    return " | ".join(dict.fromkeys(values)) or None


def _row(record: dict[str, Any], annotation: dict[str, Any] | None) -> list[Any]:
    annotation = annotation or {}
    contractors = record.get("contractors", [])
    payers = record.get("payers", [])
    field_start = _dt(record.get("field_start"))
    field_end = _dt(record.get("field_end"))
    period = None
    if field_start or field_end:
        period = f"{field_start:%Y-%m-%d}" if field_start else ""
        period += " a "
        period += f"{field_end:%Y-%m-%d}" if field_end else ""
    return [
        str(record.get("protocol") or ""),
        record.get("election_name"),
        _dt(record.get("registered_at")),
        _dt(record.get("publication_allowed_at")),
        record.get("uf"),
        record.get("geographic_level"),
        record.get("electoral_unit_name"),
        record.get("positions_raw"),
        record.get("company_name"),
        record.get("company_trade_name"),
        str(record.get("company_cnpj") or "") if record.get("company_cnpj") else None,
        _join_names(contractors, "name"),
        _join_names(payers, "name"),
        _money(record.get("research_value_cents")),
        record.get("sample_size"),
        period,
        record.get("methodology"),
        record.get("sample_plan"),
        record.get("statistician_name"),
        record.get("statistician_registration"),
        record.get("availability_status", "ativa"),
        record.get("publication_status"),
        _dt(record.get("first_seen_at")),
        _dt(record.get("last_checked_at")),
        _dt(record.get("last_changed_at")),
        _dt(record.get("actual_publication_at")),
        record.get("publication_source"),
        annotation.get("notes") or record.get("manual_notes"),
        annotation.get("tags") or record.get("manual_tags"),
        annotation.get("priority") if annotation.get("priority") is not None else record.get("manual_priority"),
        record.get("document_references", {}).get("questionario"),
        record.get("document_references", {}).get("territorio"),
        record.get("document_references", {}).get("nota_fiscal"),
        record.get("source_reference"),
        record.get("business_hash"),
    ]


def import_manual_annotations(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "Observações manuais" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Observações manuais"]
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        indices = {name: i for i, name in enumerate(headers)}
        required = {"Número do registro", "Observações", "Tags", "Prioridade"}
        if not required.issubset(indices):
            wb.close()
            raise ValueError("A aba Observações manuais não possui o cabeçalho esperado")
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            protocol = values[indices["Número do registro"]]
            if not protocol:
                continue
            rows.append({"protocol": str(protocol), "notes": values[indices["Observações"]], "tags": values[indices["Tags"]], "priority": values[indices["Prioridade"]]})
        wb.close()
        return rows
    except Exception:
        raise


def _apply_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _style_data_sheet(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_cols = {7, 8, 9, 10, 12, 13, 16, 17, 18, 27, 28, 31, 32, 33, 34}
    for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in wrap_cols)
        if any(len(str(cell.value or "")) > 55 for cell in row):
            ws.row_dimensions[row_index].height = 34
    for col in range(1, len(headers) + 1):
        max_len = min(42, max([len(str(ws.cell(r, col).value or "")) for r in range(1, min(ws.max_row, 50) + 1)] + [len(headers[col - 1])]))
        ws.column_dimensions[get_column_letter(col)].width = max(12, max_len + 2)
    for col, header in enumerate(headers, start=1):
        if "data/hora" in header.lower() or header.lower() in {"início", "fim"}:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for item in cell:
                    item.number_format = "yyyy-mm-dd hh:mm:ss"
        elif any(token in header.lower() for token in ("data permitida", "primeira localização", "última verificação", "última alteração", "publicado em", "localizado em")):
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for item in cell:
                    item.number_format = "yyyy-mm-dd"
    for cell in ws.iter_cols(min_col=14, max_col=14, min_row=2):
        for item in cell:
            item.number_format = '#,##0.00'
    for cell in ws.iter_cols(min_col=15, max_col=15, min_row=2):
        for item in cell:
            item.number_format = '#,##0'
    if rows:
        _apply_table(ws, f"A1:{get_column_letter(len(headers))}{len(rows) + 1}", table_name)


def build_workbook(records: list[dict[str, Any]], annotations: dict[str, dict[str, Any]], metrics: dict[str, Any], changes: list[dict[str, Any]], runs: list[dict[str, Any]], errors: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    panel = wb.create_sheet("Painel")
    new_ws = wb.create_sheet("Novos registros")
    upcoming_ws = wb.create_sheet("Próximas divulgações")
    base_ws = wb.create_sheet("Base completa")
    change_ws = wb.create_sheet("Alterações")
    pubs_ws = wb.create_sheet("Publicações confirmadas")
    runs_ws = wb.create_sheet("Execuções e erros")
    notes_ws = wb.create_sheet("Observações manuais")
    navy = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")
    panel.sheet_view.showGridLines = False
    panel["A1"] = "Monitor de pesquisas eleitorais — TSE"
    panel["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    panel["A1"].fill = navy
    panel.merge_cells("A1:D1")
    panel["A3"] = "Indicador"
    panel["B3"] = "Valor"
    for cell in panel[3]:
        cell.fill = navy
        cell.font = Font(color="FFFFFF", bold=True)
    labels = [
        ("Última execução", metrics.get("last_execution")),
        ("Geração da fonte", metrics.get("source_generation")),
        ("Pesquisas conhecidas", len(records)),
        ("Novas pesquisas", metrics.get("new_count", 0)),
        ("Pesquisas alteradas", metrics.get("changed_count", 0)),
        ("Divulgação permitida hoje", metrics.get("today_count", 0)),
        ("Divulgação permitida amanhã", metrics.get("tomorrow_count", 0)),
        ("Próximos sete dias", metrics.get("next_7_count", 0)),
        ("Erros da última execução", metrics.get("error_count", 0)),
    ]
    for row_index, (label, value) in enumerate(labels, start=4):
        panel.cell(row_index, 1, label)
        panel.cell(row_index, 2, value)
        panel.cell(row_index, 1).fill = light
    panel.column_dimensions["A"].width = 34
    panel.column_dimensions["B"].width = 28
    panel["A15"] = "Observação"
    panel["A15"].font = Font(bold=True)
    panel["A16"] = "Data permitida para divulgação não confirma publicação efetiva."
    panel["A16"].alignment = Alignment(wrap_text=True)
    panel.merge_cells("A16:D17")

    current_rows = [_row(r, annotations.get(r.get("protocol"))) for r in sorted(records, key=lambda x: (x.get("registered_at") or "", x.get("protocol") or ""))]
    _write_table_sheet(base_ws, HEADERS, current_rows, "BaseCompleta")
    new_protocols = {r.get("protocol") for r in metrics.get("new_records", [])}
    new_rows = [row for row, record in zip(current_rows, sorted(records, key=lambda x: (x.get("registered_at") or "", x.get("protocol") or ""))) if record.get("protocol") in new_protocols]
    _write_table_sheet(new_ws, HEADERS, new_rows, "NovosRegistros")
    today = date.today()
    upcoming = []
    for record in records:
        allowed = _dt(record.get("publication_allowed_at"))
        if allowed and today <= allowed.date() <= date.fromordinal(today.toordinal() + int(metrics.get("window_days", 7))):
            upcoming.append(record)
    upcoming.sort(key=lambda x: (x.get("publication_allowed_at") or "", x.get("protocol") or ""))
    _write_table_sheet(upcoming_ws, HEADERS, [_row(r, annotations.get(r.get("protocol"))) for r in upcoming], "ProximasDivulgacoes")
    upcoming_ws.conditional_formatting.add(f"A2:{get_column_letter(len(HEADERS))}{max(upcoming_ws.max_row, 2)}", FormulaRule(formula=["$D2=TODAY()"], fill=PatternFill("solid", fgColor="FFF2CC")))

    change_headers = ["Número do registro", "Data da alteração", "Tipo", "Campo", "Valor anterior", "Valor novo", "Fonte"]
    _write_table_sheet(change_ws, change_headers, [[c.get("protocol"), _dt(c.get("detected_at")), c.get("change_type"), c.get("field_path"), c.get("old_value"), c.get("new_value"), c.get("source")] for c in changes], "Alteracoes")
    pub_headers = ["Número do registro", "Status", "URL", "Título", "Publicado em", "Localizado em", "Confiança", "Revisão humana"]
    _write_table_sheet(pubs_ws, pub_headers, [], "PublicacoesConfirmadas")
    run_headers = ["Execução", "Início", "Fim", "Status", "Geração da fonte", "Registros", "Novos", "Alterados", "Erros", "Mensagem"]
    _write_table_sheet(runs_ws, run_headers, [[r.get("run_id"), _dt(r.get("started_at")), _dt(r.get("finished_at")), r.get("status"), r.get("source_generation"), r.get("row_count"), r.get("new_count"), r.get("changed_count"), r.get("error_count"), r.get("message")] for r in runs] + [["ERRO", _dt(e.get("created_at")), None, e.get("error_code"), None, None, None, None, 1, e.get("message")] for e in errors], "ExecucoesErros")
    notes_headers = ["Número do registro", "Observações", "Tags", "Prioridade"]
    note_rows = [[r.get("protocol"), r.get("notes"), r.get("tags"), r.get("priority")] for r in sorted(annotations.values(), key=lambda x: x.get("protocol", ""))]
    _write_table_sheet(notes_ws, notes_headers, note_rows, "ObservacoesManuais")
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 90
    return wb


def _write_table_sheet(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_data_sheet(ws, headers, rows, table_name)


def export_workbook(path: Path, records: list[dict[str, Any]], annotations: dict[str, dict[str, Any]], metrics: dict[str, Any], changes: list[dict[str, Any]], runs: list[dict[str, Any]], errors: list[dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(records, annotations, metrics, changes, runs, errors)
    fd, tmp_name = tempfile.mkstemp(prefix="monitor_pesquisas_tse_", suffix=".tmp.xlsx", dir=path.parent)
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        wb.save(tmp)
        check = load_workbook(tmp, read_only=True, data_only=False)
        expected = {"Painel", "Novos registros", "Próximas divulgações", "Base completa", "Alterações", "Publicações confirmadas", "Execuções e erros", "Observações manuais"}
        if set(check.sheetnames) != expected:
            raise ValueError(f"Abas exportadas inesperadas: {check.sheetnames}")
        check.close()
        os.replace(tmp, path)
        return path
    finally:
        if tmp.exists():
            tmp.unlink()
